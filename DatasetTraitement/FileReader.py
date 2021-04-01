# coding=UTF-8
# 暂时没用
# excel写入

import pydicom as dicom
import os
import cv2
import numpy as np
import sys
import tkinter as tk
import math
from tkinter import filedialog
import openpyxl

class FileReader:
    def __init__(self, dirname, DEBUG=False):
        self.dir_path = dirname
        self.path, self.filename = os.path.split(self.dir_path)
        print("dir_path: ", self.dir_path)
        print("path: ",self.path)
        print("filename: ", self.filename)
        self.DEBUG = DEBUG

        if self.DEBUG:
            print('path:', self.path, 'filename:', self.filename)

        self.book_name_xlsx = self.filename + ".xlsx"
        self.xlsx_path = self.path + "/" + self.book_name_xlsx
        print("xlsx_path: ", self.xlsx_path)

        self.sheet_name_xlsx = 'test_1'

        self.value = [["姓名", "性别", "年龄", "城市", "职业"],
                  ["111", "女", "66", "石家庄", "运维工程师"],
                  ["222", "男", "55", "南京", "饭店老板"],
                  ["333", "女", "27", "苏州", "保安"], ]

    # def write_excel_xlsx(self,path, sheet_name, value):
    def write_excel_xlsx(self):
        index = len(self.value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_name_xlsx
        for i in range(0, index):
            for j in range(0, len(self.value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(self.value[i][j]))
        workbook.save(self.xlsx_path)
        print("xlsx格式表格写入数据成功！")


    def read_excel_xlsx(self):
        workbook = openpyxl.load_workbook(self.xlsx_path)
        sheet = workbook[self.sheet_name_xlsx]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    dir_path = filedialog.askdirectory()

    filereader = FileReader(dir_path, True)
    filereader.write_excel_xlsx()
    filereader.read_excel_xlsx()